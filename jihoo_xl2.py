from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from subprocess import CREATE_NO_WINDOW
serv=Service()
serv.creation_flags = CREATE_NO_WINDOW   
driver = webdriver.Chrome(service = serv)


import time, datetime, os#시간,날짜,파일
import openpyxl

now= str(datetime.datetime.now())[:16]#지금 시간을 문자로 저장.15번째 칸까지 자르기
folderName= now.replace(":","_")#시간의 :을 _로 바꿈

os.mkdir(folderName)#폴더 만들기(변수 이름(폴더 이름))
key_word=["테슬라 주가","세종시","청소년 도박","중학생 마약"]
wb= openpyxl.Workbook()

for i in range(len(key_word)): # key_word의 길이 반복
    ws= wb.create_sheet()   #워크시트
    ws.title=key_word[i]
    ws.column_dimensions["A"].width = 90  #A열 너비 조절

    url = "https://search.naver.com/search.naver?where=news&ie=utf8&sm=nws_hty&query=" + key_word[i]
    driver.get(url)
    time.sleep(2)

    list_news = driver.find_element("class name", "list_news")
    news_boxes = list_news.find_elements("class name", "bx")
    
    
    for j in range(len(news_boxes)):
        driver.execute_script("arguments[0].scrollIntoView(true);", news_boxes[j])   #찾는 개체까지 스크롤 다운
        file=f"{folderName}/{i+1}_{key_word[i]}-{j+1}.png"
        news_boxes[j].screenshot(file) #스크린샷해서 저장
        
        ws.row_dimensions[j+1].height=100  #높이100
        img=openpyxl.drawing.image.Image(file) #삽입 가능하게 바꾸기 
        ws.add_image(img, f'A{j+1}') #이미지 삽입

        title=news_boxes[j].find_element("class name","news_tit") #파일 이름 저장
        print(j+1, title.text)
        
        link= title.get_attribute("href")
        ws[f'B{j+1}'].value="기사링크"
        ws[f'B{j+1}'].hyperlink=link

    print()


wb.remove(wb["Sheet"]) #필요없는 첫번째 시트 삭제
wb.save(f"{folderName}/{folderName}_결과.xlsx")
        

        
    
    




